"""
Consolidated multi-sheet analysis workbook (outputs/evaluation/analysis_report.xlsx).

Every evaluate.py run upserts its (model, quant_type) results into this single
shared workbook instead of writing separate scores_*.json / acoustics_*.json /
codebook_divergence.json / transcription.json files per run, so results across
models and quant types stay comparable in one place.

All visible sheets are laid out with metric columns grouped (stacked) under a
quant_type header band, so the same metric can be compared side by side across
quant types on one row:

  Summary      - one metric per row, one column per (model, quant_type)
  PerSample    - one row per (model, sample_id); columns grouped per quant_type:
                 aggregated per-sample stats plus baseline / quant transcript text
  LogProbs_KL  - one row per (model, sample_id, step); columns grouped per quant_type:
                 baseline and quant token ids (both LLM-vocab and codec-space) and
                 probabilities kept adjacent for direct comparison, plus per-step
                 KL/JS/argmax-match/top-k-jaccard. Row-level, not bucketed by
                 codec/frame position (codebook_id is an informational index column).
                 Only populated for models where teacher forcing is supported.
  Codec        - one row per (model, sample_id, hierarchy_level); columns grouped per
                 quant_type: codebook divergence rates from the RVQ/FSQ hierarchy
                 breakdown

The "_*Data" sheets are hidden helpers holding each table in long format (one
row per model/quant_type); they are the persisted source of truth that the wide
visible sheets are pivoted from on every write.

Writes take an exclusive file lock so concurrent evaluation jobs sharing one
workbook can't clobber each other's rows (lost update -> "only the last quant
type is in the report").
"""

from __future__ import annotations

import fcntl
import json
import math
from contextlib import contextmanager
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from decoding.alignment import teacher_forced_hierarchy_divergence
from decoding.distribution import first_token_divergence_rate as compute_first_token_divergence_rate
from evaluation.metrics import codebook_ids_for_tokens

PER_SAMPLE_SHEET = "PerSample"
LOGPROBS_SHEET = "LogProbs_KL"
CODEC_SHEET = "Codec"
SUMMARY_SHEET = "Summary"
SUMMARY_DATA_SHEET = "_SummaryData"
PER_SAMPLE_DATA_SHEET = "_PerSampleData"
LOGPROBS_DATA_SHEET = "_LogProbsData"
CODEC_DATA_SHEET = "_CodecData"

_KEY_COLS = ["model", "quant_type"]
_RUN_LABEL_COLS = ["comparison", "baseline_run", "quant_run"]

# Columns that identify a row in the wide layout; everything else becomes a
# per-quant_type column group.
_PER_SAMPLE_ID_COLS = ["model", "sample_id", "ground_truth_text"]
_LOGPROBS_ID_COLS = ["model", "sample_id", "ground_truth_text", "step", "codebook_id"]
_CODEC_ID_COLS = [
    "model",
    "sample_id",
    "ground_truth_text",
    "codec_family",
    "tokens_per_frame",
    "hierarchy_level",
]


@contextmanager
def _report_lock(report_path: Path):
    """Serialize read-modify-write cycles across concurrently running jobs."""
    report_path.parent.mkdir(parents=True, exist_ok=True)
    lock_path = report_path.with_name(report_path.name + ".lock")
    with open(lock_path, "w") as handle:
        fcntl.flock(handle, fcntl.LOCK_EX)
        try:
            yield
        finally:
            fcntl.flock(handle, fcntl.LOCK_UN)


def _comparison_label(quant_type: str) -> str:
    return f"baseline_vs_{quant_type}"


def _run_labels(quant_type: str) -> dict[str, str]:
    return {
        "comparison": _comparison_label(quant_type),
        "baseline_run": "baseline",
        "quant_run": quant_type,
    }


def _move_columns_after(df: pd.DataFrame, anchor: str, columns: list[str]) -> pd.DataFrame:
    ordered = list(df.columns)
    moved = [col for col in columns if col in ordered]
    if anchor not in ordered or not moved:
        return df
    remainder = [col for col in ordered if col not in moved]
    anchor_index = remainder.index(anchor) + 1
    return df[remainder[:anchor_index] + moved + remainder[anchor_index:]]


def _read_sheet(report_path: Path, sheet_name: str, columns: list[str]) -> pd.DataFrame:
    if not report_path.exists():
        return pd.DataFrame(columns=columns)
    try:
        return pd.read_excel(report_path, sheet_name=sheet_name)
    except (ValueError, KeyError):
        return pd.DataFrame(columns=columns)


def _read_long(report_path: Path, data_sheet: str, legacy_sheet: str, columns: list[str]) -> pd.DataFrame:
    """Read the long-format source of truth, falling back to pre-wide-layout workbooks."""
    stored = _read_sheet(report_path, data_sheet, columns)
    if not stored.empty:
        return stored
    legacy = _read_sheet(report_path, legacy_sheet, columns)
    if "quant_type" in legacy.columns:  # old workbooks kept the long rows on the visible sheet
        return legacy
    return pd.DataFrame(columns=columns)


def _upsert(existing: pd.DataFrame, new_rows: pd.DataFrame, model: str, quant_type: str) -> pd.DataFrame:
    """Drop any prior rows for this (model, quant_type) and append the fresh ones."""
    if not existing.empty:
        mask = (existing["model"] == model) & (existing["quant_type"] == quant_type)
        existing = existing.loc[~mask]
    return pd.concat([existing, new_rows], ignore_index=True)


def _summary_rows(
    model: str,
    quant_type: str,
    scores_summary: dict[str, Any],
    dist_summary: dict[str, Any],
    acoustics_summary: dict[str, Any],
    codebook_summary: dict[str, Any],
    baseline_transcription: dict[str, Any],
    quant_transcription: dict[str, Any],
    first_token_divergence_rate: float | None,
    teacher_forced_hierarchy: dict[str, Any],
) -> pd.DataFrame:
    # Load model size stats if available
    model_size_mb = None
    compression_ratio = None
    if quant_type != "none":
        try:
            from pathlib import Path
            quant_store = Path(__file__).parents[2] / "quant_models"
            # Build cache path: model_slug__quant_type
            import re
            slug = re.sub(r"[/\\:]+", "--", model.strip("/\\"))
            slug = re.sub(r"[\s.]+", "-", slug).strip("-")
            cache_dir = quant_store / f"{slug}__{quant_type}"
            stats_file = cache_dir / "quant_stats.json"
            if stats_file.exists():
                stats = json.loads(stats_file.read_text())
                model_size_mb = stats.get("quant_size_mb")
                compression_ratio = stats.get("compression_ratio")
        except Exception:
            pass  # Stats not available

    metrics: dict[str, Any] = {
        "baseline_wer": baseline_transcription.get("mean_wer"),
        "baseline_cer": baseline_transcription.get("mean_cer"),
        "quant_wer": quant_transcription.get("mean_wer"),
        "quant_cer": quant_transcription.get("mean_cer"),
        "mean_mcd": acoustics_summary.get("mean_mcd"),
        "mean_f0_frame_error": acoustics_summary.get("mean_f0_frame_error"),
        "mean_pitch_pearson_correlation": acoustics_summary.get("mean_pitch_pearson_correlation"),
        "mean_utmos_baseline": acoustics_summary.get("mean_utmos_baseline"),
        "mean_utmos_quant": acoustics_summary.get("mean_utmos_quant"),
        "utmos_degradation": acoustics_summary.get("utmos_degradation"),
        "model_size_mb": model_size_mb,
        "compression_ratio": compression_ratio,
        "first_token_divergence_rate": first_token_divergence_rate,
        "mean_final_divergence_rate": scores_summary.get("mean_final_divergence_rate"),
        "mean_prob_difference": scores_summary.get("mean_prob_difference"),
        "mean_kl_divergence": scores_summary.get("mean_kl_divergence"),
        "codebook_mean_divergence_rate": codebook_summary.get("mean_divergence_rate"),
    }
    if dist_summary:
        metrics["teacher_forced_mean_kl"] = dist_summary.get("mean_kl")
        metrics["teacher_forced_mean_js"] = dist_summary.get("mean_js")
        metrics["teacher_forced_argmax_mismatch_rate"] = dist_summary.get("argmax_mismatch_rate")
        metrics["teacher_forced_mean_nll_baseline"] = dist_summary.get("mean_nll_baseline")
        metrics["teacher_forced_mean_nll_quant"] = dist_summary.get("mean_nll_quant")
        metrics["teacher_forced_perplexity_baseline"] = dist_summary.get("perplexity_baseline")
        metrics["teacher_forced_perplexity_quant"] = dist_summary.get("perplexity_quant")
        metrics["num_samples_with_first_token_divergence"] = dist_summary.get(
            "num_samples_with_first_token_divergence"
        )
        metrics["num_samples_with_any_teacher_forced_argmax_mismatch"] = dist_summary.get(
            "num_samples_with_any_teacher_forced_argmax_mismatch"
        )
        metrics["mean_first_teacher_forced_mismatch_position"] = dist_summary.get(
            "mean_first_teacher_forced_mismatch_position"
        )
        for key, value in dist_summary.items():
            if key.startswith("mean_top") and key.endswith("_jaccard"):
                metrics[f"teacher_forced_{key}"] = value
        # Teacher-forced (deterministic argmax) mismatch counts per codebook hierarchy
        # level, distinct from the free-run `codebook_{level}_rate` below.
        metrics["teacher_forced_codebook_total_mismatched_tokens"] = teacher_forced_hierarchy.get("total_mismatched")
        metrics["teacher_forced_codebook_total_tokens"] = teacher_forced_hierarchy.get("total_tokens")
        for level, payload in (teacher_forced_hierarchy.get("by_hierarchy") or {}).items():
            metrics[f"teacher_forced_codebook_{level}_mismatched_tokens"] = payload.get("mismatched")
            metrics[f"teacher_forced_codebook_{level}_rate"] = payload.get("rate")
    for level, payload in (codebook_summary.get("by_hierarchy") or {}).items():
        metrics[f"codebook_{level}_rate"] = payload.get("mean_rate")

    return pd.DataFrame(
        {"metric": list(metrics.keys()), "model": model, "quant_type": quant_type, "value": list(metrics.values())}
    )


def _first_teacher_forced_mismatch_positions(dist_per_sample: list[dict[str, Any]]) -> list[int | None]:
    positions = []
    for entry in dist_per_sample:
        mismatches = (entry.get("per_step") or {}).get("argmax_mismatch") or []
        first = next((idx for idx, mismatch in enumerate(mismatches) if bool(mismatch)), None)
        positions.append(first)
    return positions


def _per_sample_rows(
    model: str,
    quant_type: str,
    per_sample_scores: list[dict[str, Any]],
    acoustics_per_sample: list[dict[str, Any]],
    codebook_per_sample: list[dict[str, Any]],
    baseline_samples: list[dict[str, Any]],
    quant_samples: list[dict[str, Any]],
) -> pd.DataFrame:
    acoustics_by_id = {s["sample_id"]: s for s in acoustics_per_sample}
    codebook_by_id = {s["sample_id"]: s for s in codebook_per_sample}

    rows = []
    for i, score in enumerate(per_sample_scores):
        sample_id = score["sample_id"]
        acoustics = acoustics_by_id.get(sample_id, {})
        codebook = codebook_by_id.get(sample_id, {})
        baseline_t = baseline_samples[i] if i < len(baseline_samples) else {}
        quant_t = quant_samples[i] if i < len(quant_samples) else {}
        baseline_wer = baseline_t.get("wer")
        baseline_cer = baseline_t.get("cer")
        quant_wer = quant_t.get("wer")
        quant_cer = quant_t.get("cer")

        rows.append(
            {
                "model": model,
                "quant_type": quant_type,
                **_run_labels(quant_type),
                "sample_id": sample_id,
                "ground_truth_text": score.get("text"),
                "baseline_transcript": baseline_t.get("transcript"),
                "quant_transcript": quant_t.get("transcript"),
                "baseline_wer": baseline_wer,
                "baseline_cer": baseline_cer,
                "quant_wer": quant_wer,
                "quant_cer": quant_cer,
                "wer_delta_vs_baseline": (
                    quant_wer - baseline_wer if baseline_wer is not None and quant_wer is not None else None
                ),
                "cer_delta_vs_baseline": (
                    quant_cer - baseline_cer if baseline_cer is not None and quant_cer is not None else None
                ),
                "mcd": acoustics.get("mcd"),
                "f0_frame_error": acoustics.get("f0_frame_error"),
                "pitch_pearson_correlation": acoustics.get("pitch_pearson_correlation"),
                "utmos_baseline": acoustics.get("utmos_baseline"),
                "utmos_quant": acoustics.get("utmos_quant"),
                "first_sampled_mismatch_position": score.get("first_sampled_mismatch_position"),
                "final_divergence_rate": score.get("final_divergence_rate"),
                "mean_prob_difference": score.get("mean_prob_difference"),
                "mean_kl_divergence": score.get("mean_kl_divergence"),
                "codebook_divergence_rate": codebook.get("divergence_rate"),
                "codec_family": codebook.get("codec_family"),
                "tokens_per_frame": codebook.get("tokens_per_frame"),
            }
        )
    return pd.DataFrame(rows)


def _logprobs_kl_rows(
    model: str,
    quant_type: str,
    dist_per_sample: list[dict[str, Any]],
    audio_token_start: int | None,
) -> pd.DataFrame:
    """One row per teacher-forced generation step (not bucketed by codec level).

    baseline/quant token ids and probs are kept adjacent so they're directly
    comparable: quant_* is the quant model's own actual (reference) token and
    its own free-run probability of it; baseline_* is the full-precision
    model's argmax prediction at that same step and its probability of the
    quant-chosen token (via teacher forcing on the quant model's sequence).
    nll_baseline/nll_quant are the corresponding -log(prob) values.
    """
    rows = []
    for entry in dist_per_sample:
        sample_id = entry.get("sample_id")
        text = entry.get("text")
        per_step = entry.get("per_step") or {}
        summary = entry.get("summary") or {}
        tokens_per_frame = summary.get("tokens_per_frame") or 1

        baseline_token_ids = entry.get("step_baseline_token_id") or []
        quant_token_ids = entry.get("step_quant_token_id") or []
        baseline_prob = entry.get("step_baseline_prob") or []
        quant_prob = entry.get("step_quant_prob") or []
        baseline_nll = entry.get("step_baseline_nll") or []
        quant_nll = entry.get("step_quant_nll") or []
        kl = per_step.get("kl") or []
        js = per_step.get("js") or []
        argmax_mismatch = per_step.get("argmax_mismatch") or []
        topk_key = next((k for k in per_step if k.startswith("top") and k.endswith("_jaccard")), None)
        topk_jaccard = per_step.get(topk_key) or [] if topk_key else []

        n = min(
            len(baseline_token_ids), len(quant_token_ids), len(baseline_prob), len(quant_prob),
            len(baseline_nll), len(quant_nll), len(kl), len(js), len(argmax_mismatch),
        )
        codebook_ids = codebook_ids_for_tokens(n, tokens_per_frame=tokens_per_frame)

        for step in range(n):
            baseline_llm_id = baseline_token_ids[step]
            quant_llm_id = quant_token_ids[step]
            rows.append(
                {
                    "model": model,
                    "quant_type": quant_type,
                    **_run_labels(quant_type),
                    "sample_id": sample_id,
                    "ground_truth_text": text,
                    "step": step,
                    "codebook_id": codebook_ids[step],
                    "llm_token_id_baseline": baseline_llm_id,
                    "llm_token_id_quant": quant_llm_id,
                    "audio_codec_token_id_baseline": (
                        baseline_llm_id - audio_token_start if audio_token_start is not None else None
                    ),
                    "audio_codec_token_id_quant": (
                        quant_llm_id - audio_token_start if audio_token_start is not None else None
                    ),
                    "baseline_prob": baseline_prob[step],
                    "quant_prob": quant_prob[step],
                    "nll_baseline": baseline_nll[step],
                    "nll_quant": quant_nll[step],
                    "kl": kl[step],
                    "js": js[step],
                    "argmax_match": not bool(argmax_mismatch[step]),
                    "top5_jaccard": topk_jaccard[step] if step < len(topk_jaccard) else None,
                }
            )
    return pd.DataFrame(rows)


def _codec_rows(model: str, quant_type: str, codebook_per_sample: list[dict[str, Any]]) -> pd.DataFrame:
    rows = []
    for entry in codebook_per_sample:
        for level, payload in (entry.get("by_hierarchy") or {}).items():
            rows.append(
                {
                    "model": model,
                    "quant_type": quant_type,
                    **_run_labels(quant_type),
                    "sample_id": entry.get("sample_id"),
                    "ground_truth_text": entry.get("text"),
                    "codec_family": entry.get("codec_family"),
                    "tokens_per_frame": entry.get("tokens_per_frame"),
                    "hierarchy_level": level,
                    "mismatched": payload.get("mismatched"),
                    "rate": payload.get("rate"),
                    "token_positions": ",".join(str(p) for p in payload.get("token_positions", [])),
                    "total_tokens": entry.get("total_tokens"),
                    "divergent_tokens": entry.get("divergent_tokens"),
                    "sample_divergence_rate": entry.get("divergence_rate"),
                }
            )
    return pd.DataFrame(rows)


def _pivot_by_quant(df: pd.DataFrame, id_cols: list[str]) -> pd.DataFrame:
    """Long rows -> one row per id_cols with metric columns stacked under each quant_type."""
    if df.empty or "quant_type" not in df.columns:
        return pd.DataFrame()
    frame = df.copy()
    ids = [col for col in id_cols if col in frame.columns]
    value_cols = [col for col in frame.columns if col not in ids and col not in _RUN_LABEL_COLS and col != "quant_type"]
    if not ids or not value_cols:
        return pd.DataFrame()
    frame = frame.drop_duplicates(subset=[*ids, "quant_type"], keep="last")
    quant_types = sorted(frame["quant_type"].astype(str).unique())
    frame["quant_type"] = frame["quant_type"].astype(str)

    wide = frame.set_index([*ids, "quant_type"])[value_cols].unstack("quant_type")
    wide.columns = wide.columns.swaplevel(0, 1)
    wide = wide.reindex(columns=pd.MultiIndex.from_product([quant_types, value_cols]))
    wide.columns.names = ["quant_type", "metric"]
    return wide.sort_index()


def _style_sheet(ws, header_rows: int = 1, index_cols: int = 0) -> None:
    """Bold the header band, freeze it, and auto-size columns for readability."""
    for row in range(1, header_rows + 1):
        for cell in ws[row]:
            cell.font = Font(bold=True)
    ws.freeze_panes = f"{get_column_letter(index_cols + 1)}{header_rows + 1}"
    for col_idx, col_cells in enumerate(ws.iter_cols(max_row=min(ws.max_row, 200)), start=1):
        best_len = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(best_len + 2, 10), 60)


def _write_wide_sheet(writer, wide: pd.DataFrame, sheet_name: str, id_cols: list[str]) -> None:
    if wide.empty:
        pd.DataFrame().to_excel(writer, sheet_name=sheet_name)
        return
    wide.to_excel(writer, sheet_name=sheet_name, merge_cells=True)
    # pandas emits: row1 quant_type band, row2 metric names, row3 index labels.
    _style_sheet(writer.sheets[sheet_name], header_rows=3, index_cols=wide.index.nlevels)


def _write_workbook(
    report_path: Path,
    summary_long: pd.DataFrame,
    per_sample: pd.DataFrame,
    logprobs_kl: pd.DataFrame,
    codec: pd.DataFrame,
) -> None:
    summary_wide = (
        summary_long.pivot_table(index="metric", columns=["model", "quant_type"], values="value", aggfunc="first")
        if not summary_long.empty
        else pd.DataFrame()
    )
    per_sample = _move_columns_after(per_sample, "quant_type", _RUN_LABEL_COLS)
    logprobs_kl = _move_columns_after(logprobs_kl, "quant_type", _RUN_LABEL_COLS)
    codec = _move_columns_after(codec, "quant_type", _RUN_LABEL_COLS)

    report_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(report_path, engine="openpyxl") as writer:
        summary_wide.to_excel(writer, sheet_name=SUMMARY_SHEET)
        _write_wide_sheet(writer, _pivot_by_quant(per_sample, _PER_SAMPLE_ID_COLS), PER_SAMPLE_SHEET, _PER_SAMPLE_ID_COLS)
        _write_wide_sheet(writer, _pivot_by_quant(logprobs_kl, _LOGPROBS_ID_COLS), LOGPROBS_SHEET, _LOGPROBS_ID_COLS)
        _write_wide_sheet(writer, _pivot_by_quant(codec, _CODEC_ID_COLS), CODEC_SHEET, _CODEC_ID_COLS)

        summary_long.to_excel(writer, sheet_name=SUMMARY_DATA_SHEET, index=False)
        per_sample.to_excel(writer, sheet_name=PER_SAMPLE_DATA_SHEET, index=False)
        logprobs_kl.to_excel(writer, sheet_name=LOGPROBS_DATA_SHEET, index=False)
        codec.to_excel(writer, sheet_name=CODEC_DATA_SHEET, index=False)

        writer.sheets[SUMMARY_SHEET].freeze_panes = "B4"
        for row in (1, 2):
            for cell in writer.sheets[SUMMARY_SHEET][row]:
                cell.font = Font(bold=True)
        for name in (SUMMARY_DATA_SHEET, PER_SAMPLE_DATA_SHEET, LOGPROBS_DATA_SHEET, CODEC_DATA_SHEET):
            writer.sheets[name].sheet_state = "hidden"


def update_report(
    report_path: Path,
    model: str,
    quant_type: str,
    scores_summary: dict[str, Any],
    per_sample_scores: list[dict[str, Any]],
    dist_summary: dict[str, Any],
    dist_per_sample: list[dict[str, Any]],
    acoustics_summary: dict[str, Any],
    acoustics_per_sample: list[dict[str, Any]],
    codebook_divergence: dict[str, Any],
    baseline_transcription: dict[str, Any],
    quant_transcription: dict[str, Any],
    audio_token_start: int | None = None,
) -> None:
    """Upsert one (model, quant_type) run's results into the shared report workbook."""
    codebook_summary = codebook_divergence.get("summary", {})
    codebook_per_sample = codebook_divergence.get("per_sample", [])
    ftdr = compute_first_token_divergence_rate(dist_per_sample)
    first_tf_mismatch_positions = _first_teacher_forced_mismatch_positions(dist_per_sample)
    finite_first_tf_positions = [pos for pos in first_tf_mismatch_positions if pos is not None]
    if dist_summary:
        dist_summary = dict(dist_summary)
        dist_summary["num_samples_with_first_token_divergence"] = sum(
            1 for pos in first_tf_mismatch_positions if pos == 0
        )
        dist_summary["num_samples_with_any_teacher_forced_argmax_mismatch"] = len(finite_first_tf_positions)
        dist_summary["mean_first_teacher_forced_mismatch_position"] = (
            sum(finite_first_tf_positions) / len(finite_first_tf_positions) if finite_first_tf_positions else math.nan
        )
    teacher_forced_hierarchy = teacher_forced_hierarchy_divergence(
        dist_per_sample, model_name=model, tokens_per_frame=dist_summary.get("tokens_per_frame"),
    )

    with _report_lock(report_path):
        summary_long = _upsert(
            _read_sheet(report_path, SUMMARY_DATA_SHEET, ["metric", *_KEY_COLS, "value"]),
            _summary_rows(model, quant_type, scores_summary, dist_summary, acoustics_summary,
                          codebook_summary, baseline_transcription, quant_transcription,
                          ftdr, teacher_forced_hierarchy),
            model, quant_type,
        )
        per_sample = _upsert(
            _read_long(report_path, PER_SAMPLE_DATA_SHEET, PER_SAMPLE_SHEET, [*_KEY_COLS, "sample_id"]),
            _per_sample_rows(model, quant_type, per_sample_scores, acoustics_per_sample, codebook_per_sample,
                              baseline_transcription.get("samples", []), quant_transcription.get("samples", [])),
            model, quant_type,
        )
        logprobs_kl = _upsert(
            _read_long(report_path, LOGPROBS_DATA_SHEET, LOGPROBS_SHEET, [*_KEY_COLS, "sample_id", "step"]),
            _logprobs_kl_rows(model, quant_type, dist_per_sample, audio_token_start),
            model, quant_type,
        )
        codec = _upsert(
            _read_long(report_path, CODEC_DATA_SHEET, CODEC_SHEET, [*_KEY_COLS, "sample_id", "hierarchy_level"]),
            _codec_rows(model, quant_type, codebook_per_sample),
            model, quant_type,
        )

        _write_workbook(report_path, summary_long, per_sample, logprobs_kl, codec)


def update_transcription_report(
    report_path: Path,
    model: str,
    quant_type: str,
    baseline_transcription: dict[str, Any],
    quant_transcription: dict[str, Any],
) -> None:
    """Upsert WER/CER transcription results into an existing report workbook."""
    with _report_lock(report_path):
        summary_long = _read_sheet(report_path, SUMMARY_DATA_SHEET, ["metric", *_KEY_COLS, "value"])
        per_sample = _read_long(report_path, PER_SAMPLE_DATA_SHEET, PER_SAMPLE_SHEET, [*_KEY_COLS, "sample_id"])
        logprobs_kl = _read_long(report_path, LOGPROBS_DATA_SHEET, LOGPROBS_SHEET, [*_KEY_COLS, "sample_id", "step"])
        codec = _read_long(report_path, CODEC_DATA_SHEET, CODEC_SHEET, [*_KEY_COLS, "sample_id", "hierarchy_level"])

        metric_values = {
            "baseline_wer": baseline_transcription.get("mean_wer"),
            "baseline_cer": baseline_transcription.get("mean_cer"),
            "quant_wer": quant_transcription.get("mean_wer"),
            "quant_cer": quant_transcription.get("mean_cer"),
        }
        new_summary = pd.DataFrame(
            {
                "metric": list(metric_values),
                "model": model,
                "quant_type": quant_type,
                "value": list(metric_values.values()),
            }
        )
        if not summary_long.empty:
            mask = (summary_long["model"] == model) & (summary_long["quant_type"] == quant_type) & (
                summary_long["metric"].isin(metric_values)
            )
            summary_long = summary_long.loc[~mask]
        summary_long = pd.concat([summary_long, new_summary], ignore_index=True)

        if per_sample.empty:
            raise ValueError(f"{report_path} has no {PER_SAMPLE_SHEET!r} rows to update.")

        mask = (per_sample["model"] == model) & (per_sample["quant_type"] == quant_type)
        row_indices = list(per_sample.index[mask])
        baseline_samples = baseline_transcription.get("samples", [])
        quant_samples = quant_transcription.get("samples", [])

        def column(samples: list[dict[str, Any]], key: str) -> list[Any]:
            return [samples[i].get(key) if i < len(samples) else None for i in range(len(row_indices))]

        def delta(baseline: list[Any], quant: list[Any]) -> list[Any]:
            return [q - b if b is not None and q is not None else None for b, q in zip(baseline, quant)]

        baseline_wer, baseline_cer = column(baseline_samples, "wer"), column(baseline_samples, "cer")
        quant_wer, quant_cer = column(quant_samples, "wer"), column(quant_samples, "cer")

        updates: dict[str, Any] = {
            **_run_labels(quant_type),
            "baseline_transcript": column(baseline_samples, "transcript"),
            "quant_transcript": column(quant_samples, "transcript"),
            "baseline_wer": baseline_wer,
            "baseline_cer": baseline_cer,
            "quant_wer": quant_wer,
            "quant_cer": quant_cer,
            "wer_delta_vs_baseline": delta(baseline_wer, quant_wer),
            "cer_delta_vs_baseline": delta(baseline_cer, quant_cer),
        }
        for name, values in updates.items():
            # These round-trip from Excel as all-NaN float64 when evaluate.py left them
            # blank; writing strings into that dtype raises in pandas 3.
            per_sample[name] = (
                per_sample[name].astype(object) if name in per_sample.columns else pd.Series(index=per_sample.index, dtype=object)
            )
            per_sample.loc[row_indices, name] = values

        _write_workbook(report_path, summary_long, per_sample, logprobs_kl, codec)


__all__ = ["update_report", "update_transcription_report"]
